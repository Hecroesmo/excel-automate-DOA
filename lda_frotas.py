# import pandas as pd
# import openpyxl
# from openpyxl.utils.dataframe import dataframe_to_rows
# # from sqlalchemy import create_engine
# # from sqlalchemy.engine import URL
# # import pyodbc


# def remove_unnamed_columns(df):
#     columns_to_remove = df.columns[df.columns.str.startswith(
#         'Unnamed')]
#     df = df.drop(columns=columns_to_remove)
#     return df


# def assign(df_df_demanda_carreira, df_df_lda_frotas, column_name):
#     print('assign:')
#     # Iterate over df_df_demanda_carreira DataFrame
#     for column in df_df_demanda_carreira.columns:
#         # print('column:', column)
#         if column != 'Carreiras':
#             # Get the last value in the column
#             value = df_df_demanda_carreira[column].iloc[1]
#             # print('value:', value)

#             # Check if the column name exists in df_df_lda_frotas 'CÓDIGO LINHA' column
#             # print("column in df_df_lda_frotas['CÓDIGO LINHA'].values: ",
#             #       column in df_df_lda_frotas['CÓDIGO LINHA'].values)
#             if column in df_df_lda_frotas['CÓDIGO LINHA'].values:

#                 # Get the corresponding row index in df_df_lda_frotas
#                 # print("df_df_lda_frotas.loc[df_df_lda_frotas['CÓDIGO LINHA'] == column].index[0]",
#                 #       df_df_lda_frotas.loc[df_df_lda_frotas['CÓDIGO LINHA'] == column].index[0])

#                 row_index = df_df_lda_frotas.loc[df_df_lda_frotas['CÓDIGO LINHA']
#                                                  == column].index[0]

#                 # Assign the value to the 'AUTOCARROS DISPONIBILIZADOS - 22ª F' column in df_df_lda_frotas
#                 df_df_lda_frotas.at[row_index, column_name] = value
#     return df_df_lda_frotas


# # print('Selecione o dia da semana a carregar:')
# data_path = r'C:\Users\HecroesmoWambano\Code\pandas\excel_automation'
# file_name = r'\LDA E HLA - RELATÓRIO DA FROTA PREVISTA_REALIZADA.xlsx'


# df_demanda_carreira = pd.read_excel(
#     data_path + r'\demanda_carreira_22-06-23.xlsx', 'Sheet1', )

# df_lda_frotas = pd.read_excel(
#     data_path + file_name, 'LDA- FROTA', )

# df_lda_frotas = remove_unnamed_columns(df_lda_frotas)
# df_demanda_carreira = remove_unnamed_columns(df_demanda_carreira)

# print('\n2ª Feira: 2ª F\n3ª Feira: 3ª F\n4ª Feira: 4ª F\n5ª Feira: 5ª F\n6ª Feira: 6ª F\nSábado: SAB\nDomingo: DOM: \n')
# weekday = input('Digite o dia da semana a carregar: ')

# # print(weekday)

# # print(df_lda_frotas[f'AUTOCARROS \nDISPONIBILIZADOS -{weekday}'])
# column_name = f'AUTOCARROS \nDISPONIBILIZADOS -{weekday}'

# assign(df_demanda_carreira, df_lda_frotas, column_name)

# print('After assign')
# print(df_lda_frotas[column_name])


# # Load the Excel file into a Workbook object
# wb_lda_frotas = openpyxl.load_workbook(data_path + file_name)

# # Select the sheet you want to write the DataFrame to
# sheet_name = 'LDA- FROTA'  # Replace with the actual sheet name
# ws_lda_frotas = wb_lda_frotas[sheet_name]

# # Clear the existing data in the sheet, excluding headers
# # ws_lda_frotas.delete_rows(2, ws_lda_frotas.max_row)

# # Convert the DataFrame to a list of lists
# list_lda_frotas = df_lda_frotas.values.tolist()

# # # Write the data to the sheet, starting from the second row
# # for row in list_lda_frotas:
# #     ws_lda_frotas.append(row)

# # Write the values from the modified DataFrame to the specific column
# for row in range(len(df_lda_frotas)):
#     value = df_lda_frotas.at[row, column_name]
#     cell = ws_lda_frotas[column_name].cell(row=row + 2)
#     cell.value = value

# # Save the modified Workbook back to the Excel file
# wb_lda_frotas.save(data_path + file_name)

# # Close the Workbook object
# wb_lda_frotas.close()


# # print('\n')
# # print(df_demanda_carreira)

# # df_lda_frotas.to_excel(
# #     data_path + file_name, 'LDA- FROTA')

# # print(df_demanda_carreira['008'])

# # server = 'ENBI_DTI-RDA'
# # database = 'tdmax_data'
# # username = 'enbi'
# # password = 'ENBI!456@'


# # conn_str = f'DRIVER={{SQL Server}};SERVER={server};DATABASE={database};UID={username};PWD={password}'
# # con = pyodbc.connect(conn_str)
# # sql_query = 'EXEC get_acompanhamento_operacional'
# # df = pd.read_sql_query(sql_query, con)


# # print(df[['CÓDIGO LINHA', 'AUTOCARROS \nDISPONIBILIZADOS - 2ª F']])
# # print(bd_df['AUTOCARROS \nDISPONIBILIZADOS - 2ª F'])


# # plusOneToColumn(df, 'Qtd. Sem Prestação de Conta')
# # print(df)
# # df.to_excel(file_path, index=False)
# # con.close()

import pandas as pd
import openpyxl


def remove_unnamed_columns(df):
    columns_to_remove = df.columns[df.columns.str.startswith('Unnamed')]
    df = df.drop(columns=columns_to_remove)
    return df


def assign(df_demanda_carreira, df_lda_frotas, column_name):
    for column in df_demanda_carreira.columns:
        if column != 'Carreiras':
            value = df_demanda_carreira[column].iloc[1]
            if column in df_lda_frotas['CÓDIGO LINHA'].values:
                row_index = df_lda_frotas.loc[df_lda_frotas['CÓDIGO LINHA']
                                              == column].index[0]
                df_lda_frotas.at[row_index, column_name] = value
    return df_lda_frotas


data_path = r'C:\Users\HecroesmoWambano\Code\pandas\excel_automation'
file_name = r'\LDA E HLA - RELATÓRIO DA FROTA PREVISTA_REALIZADA.xlsx'

df_demanda_carreira = pd.read_excel(
    data_path + r'\demanda_carreira_22-06-23.xlsx', 'Sheet1')
df_lda_frotas = pd.read_excel(data_path + file_name, 'LDA- FROTA')

df_lda_frotas = remove_unnamed_columns(df_lda_frotas)
df_demanda_carreira = remove_unnamed_columns(df_demanda_carreira)

print('\n2ª Feira: 2ª F\n3ª Feira: 3ª F\n4ª Feira: 4ª F\n5ª Feira: 5ª F\n6ª Feira: 6ª F\nSábado: SAB\nDomingo: DOM: \n')
weekday = input('Digite o dia da semana a carregar: ')
column_name = f'AUTOCARROS DISPONIBILIZADOS -{weekday}'

df_lda_frotas = assign(df_demanda_carreira, df_lda_frotas, column_name)

# Load the Excel file into a Workbook object
wb_lda_frotas = openpyxl.load_workbook(data_path + file_name)

# Select the sheet you want to write the DataFrame to
sheet_name = 'LDA- FROTA'  # Replace with the actual sheet name
ws_lda_frotas = wb_lda_frotas[sheet_name]

# Get the column index of the specified column name
column_index = None
for column in ws_lda_frotas.iter_cols(min_row=1, max_row=1):
    if column[0].value == column_name:
        column_index = column[0].column
        break

# Write the values from the modified DataFrame to the specific column
for row, value in enumerate(df_lda_frotas[column_name]):
    cell = ws_lda_frotas.cell(row=row + 2, column=column_index)
    cell.value = value

# Save the modified Workbook back to the Excel file
wb_lda_frotas.save(data_path + file_name)

# Close the Workbook object
wb_lda_frotas.close()
