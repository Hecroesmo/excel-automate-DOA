import pandas as pd
import openpyxl


def remove_unnamed_columns(df):
    columns_to_remove = df.columns[df.columns.str.startswith('Unnamed')]
    df = df.drop(columns=columns_to_remove)
    return df


def assign(df_demanda_carreira, df_viagens, column_name):
    for column in df_demanda_carreira.columns:
        if column != 'Carreiras':
            value = df_demanda_carreira[column].iloc[0]
            if column in df_viagens['CÓDIGO LINHA'].values:
                row_index = df_viagens.loc[df_viagens['CÓDIGO LINHA']
                                           == column].index[0]
                df_viagens.at[row_index, column_name] = value
    return df_viagens


data_path = r'C:\Users\HecroesmoWambano\Code\pandas\excel_automation'
file_name = r'\LDA E HLA - RELATÓRIO DA FROTA PREVISTA_REALIZADA.xlsx'

df_demanda_carreira = pd.read_excel(
    data_path + r'\demanda_carreira_22-06-23.xlsx', 'Sheet1')
df_viagens = pd.read_excel(data_path + file_name, 'LDA-VIAGENS')

df_viagens = remove_unnamed_columns(df_viagens)
df_demanda_carreira = remove_unnamed_columns(df_demanda_carreira)

print('\n2ª Feira: 2ª F\n3ª Feira: 3ª F\n4ª Feira: 4ª F\n5ª Feira: 5ª F\n6ª Feira: 6ª F\nSábado: SAB\nDomingo: DOM: \n')
weekday = input('Digite o dia da semana a carregar: ')
column_name = f'VIAGENS REALIZADAS -{weekday}'

print('before assign: ')
print(df_viagens[column_name])

df_viagens = assign(df_demanda_carreira, df_viagens, column_name)

print('after assign: ')
print(df_viagens[column_name])

# Load the Excel file into a Workbook object
wb_viagens = openpyxl.load_workbook(data_path + file_name)

# Select the sheet you want to write the DataFrame to
sheet_name = 'LDA-VIAGENS'
ws_viagens = wb_viagens[sheet_name]

# Get the column index of the specified column name
column_index = None
for column in ws_viagens.iter_cols(min_row=1, max_row=1):
    if column[0].value == column_name:
        column_index = column[0].column
        break

# Write the values from the modified DataFrame to the specific column
for row, value in enumerate(df_viagens[column_name]):
    cell = ws_viagens.cell(row=row + 2, column=column_index)
    cell.value = value

# Save the modified Workbook back to the Excel file
wb_viagens.save(data_path + file_name)

# Close the Workbook object
wb_viagens.close()
