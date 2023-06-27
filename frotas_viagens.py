import pandas as pd
import openpyxl


def remove_unnamed_columns(df):
    columns_to_remove = df.columns[df.columns.str.startswith('Unnamed')]
    df = df.drop(columns=columns_to_remove)
    return df


def assign(df_demanda_carreira, def_relatorio, column_name, index):
    for column in df_demanda_carreira.columns:
        if column != 'Carreiras':
            value = df_demanda_carreira[column].iloc[index]
            if column in def_relatorio['CÓDIGO LINHA'].values:
                row_index = def_relatorio.loc[def_relatorio['CÓDIGO LINHA']
                                              == column].index[0]
                def_relatorio.at[row_index, column_name] = value
    return def_relatorio


print('\n2ª Feira: 2ª F\n3ª Feira: 3ª F\n4ª Feira: 4ª F\n5ª Feira: 5ª F\n6ª Feira: 6ª F\nSábado: SAB\nDomingo: DOM: \n')
weekday = input('Digite o dia da semana a carregar: ')

print('\nFrotas: 1\nViagens: 2\n')
option = int(input('Pretende carregar viagens ou frotas: '))

if option == 1:
    # Select the sheet you want to write the DataFrame to
    sheet_name = 'LDA- FROTA'
    column_name = f'AUTOCARROS DISPONIBILIZADOS -{weekday}'
    index = 1
else:
    # Select the sheet you want to write the DataFrame to
    sheet_name = 'LDA-VIAGENS'
    column_name = f'VIAGENS REALIZADAS -{weekday}'
    index = 0


# data_path = r'C:\Users\HecroesmoWambano\Code\pandas\excel_automation'
data_path = '/app/excel_automation'
# file_name = r'\relatorio.xlsx'
file_name = '/relatorio.xlsx'
sheet_name_list = ['LDA- FROTA', 'LDA-VIAGENS']

# df_demanda_carreira = pd.read_excel(
#     data_path + r'\demanda_carreira.xlsx', 'Sheet1')
df_demanda_carreira = pd.read_excel(
    data_path + '/demanda_carreira.xlsx', 'Sheet1')
def_relatorio = pd.read_excel(data_path + file_name, sheet_name)

def_relatorio = remove_unnamed_columns(def_relatorio)
df_demanda_carreira = remove_unnamed_columns(df_demanda_carreira)

def_relatorio = assign(df_demanda_carreira, def_relatorio, column_name, index)

# Load the Excel file into a Workbook object
wb_relatorio = openpyxl.load_workbook(data_path + file_name)

ws_relatorio = wb_relatorio[sheet_name]

# Get the column index of the specified column name
column_index = None
for column in ws_relatorio.iter_cols(min_row=1, max_row=1):
    if column[0].value == column_name:
        column_index = column[0].column
        break

# Write the values from the modified DataFrame to the specific column
for row, value in enumerate(def_relatorio[column_name]):
    cell = ws_relatorio.cell(row=row + 2, column=column_index)
    cell.value = value

# Save the modified Workbook back to the Excel file
wb_relatorio.save(data_path + file_name)

# Close the Workbook object
wb_relatorio.close()
print(f'Dados de {sheet_name_list[option - 1]} carregado com sucesso!')
